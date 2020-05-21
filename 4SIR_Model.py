###############################################################################
########################   PROJECT PROGRAMMING ################################
###############################################################################


######################         IMPORT DATA      ###############################

import pandas as pd
import numpy as np
from scipy.integrate import odeint
import matplotlib.pyplot as plt
import openpyxl as xl

#import data of the number of deaths 

path_deaths = "Number_Of_Deaths_Final.xlsx"
deaths = pd.read_excel(path_deaths, sheet_name = "By_country")



# import data on the population by country

path_population = "Population_By_Country_Final.xlsx"
population = pd.read_excel(path_population, sheet_name = "By_country")


 
# countries that are in the data base 
countries = population.Country.unique()



# Create the variables that will predict the growth rate of the population
# In the range of ages that will be analysed

for country in countries:
    globals()['growth_rate_population_'+str(country[1:]+"_Youngs")]= 0
    globals()['growth_rate_population_'+str(country[1:]+"_Adults")]= 0
    globals()['growth_rate_population_'+str(country[1:]+"_Elderly")]= 0
    
    
# Create the variables that will predict the growth rate of the number of deaths
# In the range of ages that will be analysed

for country in countries:
    globals()['growth_rate_deaths_'+str(country[1:]+"_Youngs")] = 0
    globals()['growth_rate_deaths_'+str(country[1:]+"_Adults")] = 0
    globals()['growth_rate_deaths_'+str(country[1:]+"_Elderly")] = 0


# Function that will compute the growth rates 
def compute_growth_rates():

   
    # Take the population for each country separatly
    
    for country in countries:
        globals()['population_'+ str(country[1:])] = \
                              population.loc[population['Country'] == country]
    
    # Take the number of deaths for each country separatly
        
    for country in countries:
        globals()['deaths_'+ str(country[1:])] = \
                                       deaths.loc[deaths['Country'] == country]
    
    
    # We are interested on the columns with the numbers to compute growth rates 
    my_cols = set(deaths.columns)
    my_cols.remove("KEY")
    my_cols.remove("Country")
    
    
    
    # Computes the percentage change on the number of deaths 
    
    for country in countries:
        globals()['pctChange_deaths_'+ str(country[1:])] = \
                    globals()['deaths_'+str(country[1:])][my_cols].pct_change()
    
   
    # Computes the mean of the percentage change on the number of deaths
                    
    for country in countries:
        globals()['mean_pctChange_deaths_'+ str(country[1:])] = \
                         globals()['pctChange_deaths_'+str(country[1:])].mean()  
    
    
    
    # Computes the percentage change on the population 
                         
    for country in countries:
        globals()['pctChange_population_'+ str(country[1:])] = \
                globals()['population_'+str(country[1:])][my_cols].pct_change()
        
        
    # Computes the mean of the percentage change on the population 
                
    for country in countries:
        globals()['mean_pctChange_population_'+ str(country[1:])] = \
                    globals()['pctChange_population_'+str(country[1:])].mean()  
    
    
    # Growth rate of the population on the range of ages to analyze
                    
    for country in countries:
        globals()['growth_rate_population_'+str(country[1:]+"_Youngs")] = \
            globals()['mean_pctChange_population_'+str(country[1:])]["Youngs"]
            
        globals()['growth_rate_population_'+str(country[1:]+"_Adults")] = \
            globals()['mean_pctChange_population_'+str(country[1:])]["Adults"]
        
        globals()['growth_rate_population_'+str(country[1:]+"_Elderly")] = \
            globals()['mean_pctChange_population_'+str(country[1:])]["Elderly"]
    

    # Growth rate of the number of deaths on the range of ages to analyze
        
    for country in countries:
        globals()['growth_rate_deaths_'+str(country[1:]+"_Youngs")] = \
                globals()['mean_pctChange_deaths_'+str(country[1:])]["Youngs"]
                
        globals()['growth_rate_deaths_'+str(country[1:]+"_Adults")] = \
                globals()['mean_pctChange_deaths_'+str(country[1:])]["Adults"]
                
        globals()['growth_rate_deaths_'+str(country[1:]+"_Elderly")] = \
                globals()['mean_pctChange_deaths_'+str(country[1:])]["Elderly"]
    
    ## delete the variables that we don't need anymore 
                 
    for country in countries:
        del globals()['pctChange_deaths_'+ str(country[1:])] 
    
    for country in countries:
        del globals()['mean_pctChange_deaths_'+ str(country[1:])]
     
    for country in countries:
        del globals()['pctChange_population_'+ str(country[1:])]       

    for country in countries:
        del globals()['mean_pctChange_population_'+ str(country[1:])]
        
    
    
compute_growth_rates()



# Gives a Dataframe with the population of each range of age by country

for country in countries:
    globals()['population_'+ str(country[1:])+'_Youngs'] = \
                 globals()['population_'+ str(country[1:])][["Year","Youngs"]]
                 
    globals()['population_'+ str(country[1:])+'_Adults'] = \
                 globals()['population_'+ str(country[1:])][["Year","Adults"]]
                 
    globals()['population_'+ str(country[1:])+'_Elderly'] = \
                globals()['population_'+ str(country[1:])][["Year","Elderly"]]




# Gives a Dataframe with the number of deaths of each range of age by country
    
for country in countries:
    globals()['deaths_'+ str(country[1:])+'_Youngs'] = \
                    globals()['deaths_'+ str(country[1:])][["Year","Youngs"]]
                    
    globals()['deaths_'+ str(country[1:])+'_Adults'] = \
                    globals()['deaths_'+ str(country[1:])][["Year","Adults"]]
                    
    globals()['deaths_'+ str(country[1:])+'_Elderly'] = \
                    globals()['deaths_'+ str(country[1:])][["Year","Elderly"]]



# Projections of the population in the future years assuming that it continues
# with the same growth rate

for country in countries:
    year_max = globals()['population_'+ str(country[1:])+'_Youngs']["Year"].iloc[-1]
    for ages in ["Youngs","Adults","Elderly"]:
        year = 1
        while (year_max+year) <= 2030:
            
            # The value that we have to add to the dataframe
            additional = \
                globals()['population_'+ str(country[1:])+'_'+ ages][ages].iloc[-1]\
                *((1+globals()['growth_rate_population_'+str(country[1:]+"_"+ages)]))
            
            # Create a dataframe with the adequate year to be appended
            toapend = pd.DataFrame([[year_max + year,additional]], \
                                                         columns=["Year",ages])
            
            # Append it to the population dataframe
            globals()['population_'+ str(country[1:])+'_'+ages] = \
                globals()['population_'+ str(country[1:])+'_'+ages]\
                                         .append(toapend, ignore_index = True)
            
            year = year + 1
            del additional
            del toapend




# Projections of the number of deaths in the future years assuming that it 
#continues with the same rate

for country in countries:
    year_max = globals()['deaths_'+ str(country[1:])+'_Youngs']["Year"].iloc[-1]
    for ages in ["Youngs","Adults","Elderly"]:
        year = 1
        # To have the population up to 2030
        while (year_max+year) <= 2030:
            
            # The value that we have to add to the dataframe
            additional = \
                globals()['deaths_'+ str(country[1:])+'_'+ ages][ages].iloc[-1]\
                *((1+globals()['growth_rate_deaths_'+str(country[1:]+"_"+ages)]))
            
            # Create a dataframe with the adequate year to be appended
            toapend = pd.DataFrame([[year_max + year,additional]], \
                                                         columns=["Year",ages])
            
            # append it to the number of deaths dataframe
            globals()['deaths_'+ str(country[1:])+'_'+ages] = \
                globals()['deaths_'+ str(country[1:])+'_'+ages].append(toapend\
                                                        , ignore_index = True)
            
            year = year + 1
            del additional
            del toapend






## Since we don't have the date on the same dates and the same indexes we 
## equalize it
            
def equalize_data():
    for country in countries:
        for ages in ["Youngs","Adults","Elderly"]:
            
            ## Take the first year of the data of the number of deaths
            globals()["min_deaths_"+ str(country[1:])+'_'+ages] = \
            globals()['deaths_'+ str(country[1:])+'_'+ ages]["Year"].iloc[0]
            
            ## Take the first year of the data of the population
            globals()["min_population_"+ str(country[1:])+'_'+ages] = \
            globals()['population_'+ str(country[1:])+'_'+ ages]["Year"].iloc[0]
                
            
            ## Start the comparable data from the biggest of the two 
            if globals()["min_deaths_"+ str(country[1:])+'_'+ages] < \
                globals()["min_population_"+ str(country[1:])+'_'+ages]:
                    
                minimum_year = globals()["min_population_"+ str(country[1:])+'_'+ages]
            else:
                minimum_year = globals()["min_deaths_"+ str(country[1:])+'_'+ages]
            
            
            ## Only take the values above the minimum year and assign it to 
            ## the comparable population 
                
            globals()['compar_population_'+ str(country[1:])+'_'+ ages]=\
                globals()['population_'+ str(country[1:])+'_'+ ages].\
                    loc[globals()['population_'+ str(country[1:])+'_'+\
                                     ages]["Year"]>= minimum_year]
            
            
            ## Same for the number of deaths 
                        
            globals()['compar_deaths_'+ str(country[1:])+'_'+ ages]= \
                      globals()['deaths_'+ str(country[1:])+'_'+ ages].\
                      loc[globals()['deaths_'+ str(country[1:])+'_'+ \
                                    ages]["Year"]>= minimum_year]
            
            
            
            # in order to have the right indexes, we reindex the comparable
            # Data_Frames equally
                          
            number = globals()['compar_population_'+ str(country[1:])+'_'+\
                                                               ages].shape[0]
            newlist = []
            i = 0
            for i in range(number):
                newlist.append(i)
            globals()['compar_population_'+ str(country[1:])+'_'+ ages] =\
                globals()['compar_population_'+ str(country[1:])+'_'\
                                                  + ages].set_index([newlist])
                    
            globals()['compar_deaths_'+ str(country[1:])+'_'+ ages] =\
                globals()['compar_deaths_'+ str(country[1:])+'_'+ \
                                                     ages].set_index([newlist])
            
            ## Delete useless variables 
            del minimum_year
            del globals()["min_deaths_"+ str(country[1:])+'_'+ages] 
            del globals()["min_population_"+ str(country[1:])+'_'+ages]
            del newlist
            del number
        

equalize_data()





## MORTALITY RATES 

## compute mmortality rates by division on the number of deaths and the population

for country in countries:
    for ages in ["Youngs","Adults","Elderly"]:
        year = globals()['compar_deaths_'+ str(country[1:])+'_'+ ages]["Year"]
        mortality = pd.DataFrame((globals()['compar_deaths_'+ \
                    str(country[1:])+'_'+ ages][ages])/(globals()\
                    ['compar_population_'+ str(country[1:])+'_'+ ages][ages]))
                                                        
        globals()["mortality_rate_"+ str(country[1:])+'_'+ ages] = \
            pd.concat([year,mortality],axis=1)
            
        del year
        del mortality



## The model that will predict the number of people suceptible of being infected
## The number of people recovered and the number of infected people given some
## arguments:
            ## POPULATION: The population of the country to analyse 
            ## CONTACT_RATE: Rate at which an infective will enconter a 
            ##               susceptible
            ## RECOVERY_RATE: Rate at which the infectives will recover from
            ##                the virus
            ## DEATH_RATE : We assume that a percentage of the recovered people
            ##              is in fact dead.
            ## DAYS : How long do we want to see the evolution
            ## FACTOR : Factor to normalize the graphs given 
            ## SCALE : To addapt the scale of the graph according to the data
            ## GOV_AJUST_RATE : After some time the government will take some
            ##                  initiatives that will reduce the contact rate
            ##                  by this factor
            ## PEOP_PREC : After some time people will understand that they
            ##             have to be more carefull and it will reduce the 
            ##             contact rate

def SIR_MODEL(COUNTRY,POPULATION,CONTACT_RATE,RECOVERY_RATE,DEATH_RATE,DAYS,FACTOR,\
              SCALE,GOV_ADJUST_RATE = 1,PEOP_PREC = 1):
    
    # Total population of the country ( N in the function )
    population_country = POPULATION

    # Initial number of infected ( I in the function )
    initial_infected = 1 

    # Initial number of recovered individuals.  ( R in the function)
    inital_recovered = 0 
    
    # Everyone else is susceptible to infection initially. ( S in the function)
    initial_susceptible = population_country - initial_infected - inital_recovered

    # Contact rate ( beta in the function) --> (in 1/days)
    contact_rate = CONTACT_RATE
    
    # mean recovery rate ( gamma in the function ) --> (in 1/days)
    recovery_rate = RECOVERY_RATE
    
    # R0  is the number of people infected by 1 single one
    # Important measure since it's the one that allow us to connect our model to 
    # COVID-19 since we have an estimation of it arround 1.5%

    R0 = (contact_rate)/recovery_rate
    print("The R0 of this model is:", R0)

    # A grid of time points (in days)
    t = np.linspace(1, DAYS, DAYS)

    
    # This will take into account any action from the government in order to 
    # Reduce the contact_rate
    
    def adjust_contact_rate(rate,day):
        if day < 30:
            contact_rate = rate
        elif day > 180:
            contact_rate = rate*GOV_ADJUST_RATE*PEOP_PREC
        else:
            contact_rate = rate*GOV_ADJUST_RATE
        
        return contact_rate
 
    

    # Equation of differentials that are the base of our Model    
    def differentials(y, t, N, beta, gamma):
    
        S, I, R = y
        
        #Will take into account actions from the state
        beta = adjust_contact_rate(beta,t)
        
        #Rate of change of the number of subceptibles
        dSdt = -beta * S * I / N
        
        # Rate of change of the infectives
        dIdt = beta * S * I / N - gamma * I
        
        # Rate of change of the number of Recovered
        dRdt = gamma * I
        
        return dSdt, dIdt, dRdt


    
    # Initial conditions vector
    initial_values = initial_susceptible, initial_infected, inital_recovered
    
    
    # Integrate the SIR equations over the time grid, t.
    ret = odeint(differentials, initial_values, t, args=(population_country,\
                                                contact_rate, recovery_rate))
    
    
    # Gives the values of each of them in an array 
    Susceptible, Infected, Recovered = ret.T
    
    
    # Plot the data on three separate curves for S(t), I(t) and R(t)
    
    fig = plt.figure(facecolor='w')
    ax = fig.add_subplot(111, facecolor='#dddddd', axisbelow=True)
    ax.plot(t, Susceptible/FACTOR, 'b', alpha=0.5, lw=2, label='Susceptible')
    ax.plot(t, Infected/FACTOR, 'r', alpha=0.5, lw=2, label='Infected')
    ax.plot(t, Recovered/FACTOR, 'g', alpha=0.5, lw=2, label='Recovered')
    ax.set_xlabel('Time/days')
    ax.set_ylabel('Number ('+str(FACTOR)+'s)')
    ax.set_ylim(0,SCALE)
    ax.xaxis.set_tick_params(length=0)
    legend = ax.legend()
    legend.get_frame().set_alpha(0.5)
    plt.title('R0 for the country: ' + COUNTRY + ' estimated at: %f ' %R0)
    plt.savefig('plots/SIR-' + COUNTRY + '.png')
    plt.show()
    
    
    ## Since the model doesn't predict any deaths but only the infected,
    ## recovered and subsceptiblels we use in our case the fact that those 
    ## who are recovered at the end of the period means that went out from the
    ## deasise.  Here we will make the assumption that from these recovered 
    ## people a percentage is in fact dead.

    
    number_of_deaths_2020 = Recovered[-1]*DEATH_RATE
    
    return number_of_deaths_2020




## From the number of deaths we will separate the proportion with these rates 
## found in the website 
    

rate_Youngs = 0.0004 + (30-18)*(0.045/(44-18))
rate_Adults = (44-30)*(0.045/(44-18)) + (60-45)*(0.231/(64-45))
rate_Elderly = (64-60)*(0.231/(64-45)) + 0.246 + 0.477

print(rate_Youngs)
print(rate_Adults)
print(rate_Elderly)


## We will only look into 5 countries

# Taking constant the recovery rate which will be 7 days to recover from the 
# infection since after that you then know that you are infected and isolate 
## yourself from the others.
# Then we change the contact_rate order to fit the best possible each country. 
# Assuming that the measures taken by the state reduced by 50% the interactions
# between the population.
# Since wehave the actual number of deaths we approximate the contact_rate by
# running a model that isn't far from the true value.  


# These are the values for the different countries
    
##############################################################################
######################     BELGIUM     #######################################
##############################################################################

## The true number of deaths is arround 8000 for the first 180 days 
## Then the contact rate that approximate the good enough is 1/2.99
## After some tests we found out that 
# this modeel : SIR_MODEL(Population_Belgium_2020,1/2.99,1/7,0.15,180,1000,35,0.5,0.6)
# is the one whoe fits the best the belgium case 
## And a plausible modeling for the whole year would be :
# SIR_MODEL(Population_Belgium_2020,1/2.99,1/7,0.15,360,1000,105,0.5,0.6), the 
# 0.6 at the end means that after some point here half year people understand,
# the gravity of the situation and have more precautions in their day life. 


# SIR_MODEL(11500000,1/2.99,1/7,0.15,180,1000,35,0.5,0.6)
# SIR_MODEL(11500000,1/2.99,1/7,0.15,360,1000,105,0.5,0.6)


contact_rate_Belgium = 1/2.99

total_population_Belgium_2020 = 0

def funct_totalpopulationBelgium():
    for ages in ["Youngs","Adults","Elderly"]:
        globals()["Deaths_2020_"+str(ages)] =\
             globals()["compar_population_Belgium_"+str(ages)][ages].loc[globals()\
             ["compar_population_Belgium_"+str(ages)]["Year"]==2020]
    total_population = globals()["Deaths_2020_Youngs"]\
                       + globals()["Deaths_2020_Adults"] \
                       + globals()["Deaths_2020_Elderly"]
    return total_population


Population_Belgium_2020 = funct_totalpopulationBelgium()

total_deaths_Belgium_2020= SIR_MODEL('Belgium',Population_Belgium_2020,\
                                     contact_rate_Belgium,\
                                     1/7,0.15,360,1000,105,0.5,0.6)


nr_deaths_Belgium_Youngs = rate_Youngs *  total_deaths_Belgium_2020
nr_deaths_Belgium_Adults = rate_Adults * total_deaths_Belgium_2020
nr_deaths_Belgium_Elderly = rate_Elderly * total_deaths_Belgium_2020

print(nr_deaths_Belgium_Youngs)
print(nr_deaths_Belgium_Adults)
print(nr_deaths_Belgium_Elderly)


## We make a copy from the number of deaths and we assume that the only year 
## that it changes is in 2020 the rest stays constant 

for ages in ["Youngs","Adults","Elderly"]:
    globals()["covid_deaths_Belgium_"+str(ages)] = \
        globals()["compar_deaths_Belgium_" + str(ages)].copy()
        
    globals()["covid_population_Belgium_"+str(ages)] = \
        globals()["compar_population_Belgium_" + str(ages)].copy()

    
    globals()["covid_deaths_Belgium_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Belgium_"+str(ages)]["Year"]==2020] =\
             globals()["covid_deaths_Belgium_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Belgium_"+str(ages)]["Year"]==2020]  \
             + globals()["nr_deaths_Belgium_"+ str(ages)]
            
            
    globals()["covid_population_Belgium_"+str(ages)][ages].loc[globals()\
             ["covid_population_Belgium_"+str(ages)]["Year"]==2020] =\
        globals()["covid_population_Belgium_"+str(ages)][ages].loc[globals()\
        ["covid_population_Belgium_"+str(ages)]["Year"]==2020] +\
            globals()["nr_deaths_Belgium_"+ str(ages)]
    
    
    year_Belgium = globals()["covid_population_Belgium_Youngs"]["Year"]
    
    globals()["mortality_Belgium_" + str(ages)] = pd.DataFrame( \
                     (globals()["covid_deaths_Belgium_"+str(ages)][ages])/\
                    (globals()["covid_population_Belgium_"+str(ages)][ages]))
    
    globals()["covid_mortality_rate_Belgium_" + str(ages)] = \
                     pd.concat([year_Belgium,globals()["mortality_Belgium_" +\
                                                       str(ages)]],axis = 1)
    
    del year_Belgium
    del globals()["mortality_Belgium_" + str(ages)]


Deaths_covid_Belgium_Youngs = globals()["covid_deaths_Belgium_Youngs"].loc[globals()["covid_deaths_Belgium_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_withoutcovid_Belgium_Youngs = globals()["compar_deaths_Belgium_Youngs"].loc[globals()["covid_deaths_Belgium_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_covid_Belgium_Adults = globals()["covid_deaths_Belgium_Adults"].loc[globals()["covid_deaths_Belgium_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_withoutcovid_Belgium_Adults = globals()["compar_deaths_Belgium_Adults"].loc[globals()["covid_deaths_Belgium_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_covid_Belgium_Elderly = globals()["covid_deaths_Belgium_Elderly"].loc[globals()["covid_deaths_Belgium_Elderly"]["Year"]==2020]["Elderly"].values[0]
Deaths_withoutcovid_Belgium_Elderly = globals()["compar_deaths_Belgium_Elderly"].loc[globals()["covid_deaths_Belgium_Elderly"]["Year"]==2020]["Elderly"].values[0]


mortality_covid_Belgium_Youngs = globals()["covid_mortality_rate_Belgium_Youngs"].loc[globals()["covid_mortality_rate_Belgium_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_withoutcovid_Belgium_Youngs = globals()["mortality_rate_Belgium_Youngs"].loc[globals()["mortality_rate_Belgium_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_covid_Belgium_Adults = globals()["covid_mortality_rate_Belgium_Adults"].loc[globals()["covid_mortality_rate_Belgium_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_withoutcovid_Belgium_Adults = globals()["mortality_rate_Belgium_Adults"].loc[globals()["mortality_rate_Belgium_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_covid_Belgium_Elderly = globals()["covid_mortality_rate_Belgium_Elderly"].loc[globals()["covid_mortality_rate_Belgium_Elderly"]["Year"]==2020]["Elderly"].values[0]
mortality_withoutcovid_Belgium_Elderly = globals()["mortality_rate_Belgium_Elderly"].loc[globals()["mortality_rate_Belgium_Elderly"]["Year"]==2020]["Elderly"].values[0]


print((Deaths_covid_Belgium_Elderly - Deaths_withoutcovid_Belgium_Elderly)/Deaths_withoutcovid_Belgium_Elderly)




##############################################################################
######################      FRANCE     #######################################
##############################################################################




# SIR_MODEL(67500000,1/2.88,1/7,0.15,180,1000,35,0.5,0.6)
# SIR_MODEL(67500000,1/2.88,1/7,0.15,1160,1000,105,0.5,0.6)


contact_rate_France = 1/2.88 

total_population_France_2020 = 0

def funct_totalpopulationFrance():
    for ages in ["Youngs","Adults","Elderly"]:
        globals()["Deaths_2020_"+str(ages)] =\
             globals()["compar_population_France_"+str(ages)][ages].loc[globals()\
             ["compar_population_France_"+str(ages)]["Year"]==2020]
    total_population = globals()["Deaths_2020_Youngs"]\
                       + globals()["Deaths_2020_Adults"] \
                       + globals()["Deaths_2020_Elderly"]
    return total_population


Population_France_2020 = funct_totalpopulationFrance()

total_deaths_France_2020= SIR_MODEL('France',Population_France_2020,\
                                     contact_rate_France,\
                                     1/7,0.15,360,1000,105,0.5,0.6)


nr_deaths_France_Youngs = rate_Youngs *  total_deaths_France_2020
nr_deaths_France_Adults = rate_Adults * total_deaths_France_2020
nr_deaths_France_Elderly = rate_Elderly * total_deaths_France_2020



## We make a copy from the number of deaths and we assume that the only year 
## that it changes is in 2020 the rest stays constant 

for ages in ["Youngs","Adults","Elderly"]:
    globals()["covid_deaths_France_"+str(ages)] = \
        globals()["compar_deaths_France_" + str(ages)].copy()
        
    globals()["covid_population_France_"+str(ages)] = \
        globals()["compar_population_France_" + str(ages)].copy()

    
    globals()["covid_deaths_France_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_France_"+str(ages)]["Year"]==2020] =\
             globals()["covid_deaths_France_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_France_"+str(ages)]["Year"]==2020]  \
             + globals()["nr_deaths_France_"+ str(ages)]
            
            
    globals()["covid_population_France_"+str(ages)][ages].loc[globals()\
             ["covid_population_France_"+str(ages)]["Year"]==2020] =\
        globals()["covid_population_France_"+str(ages)][ages].loc[globals()\
        ["covid_population_France_"+str(ages)]["Year"]==2020] +\
            globals()["nr_deaths_France_"+ str(ages)]
    
    
    year_France = globals()["covid_population_France_Youngs"]["Year"]
    
    globals()["mortality_France_" + str(ages)] = pd.DataFrame( \
                     (globals()["covid_deaths_France_"+str(ages)][ages])/\
                    (globals()["covid_population_France_"+str(ages)][ages]))
    
    globals()["covid_mortality_rate_France_" + str(ages)] = \
                     pd.concat([year_France,globals()["mortality_France_" +\
                                                       str(ages)]],axis = 1)
    del year_France
    del globals()["mortality_France_" + str(ages)]



Deaths_covid_France_Youngs = globals()["covid_deaths_France_Youngs"].loc[globals()["covid_deaths_France_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_withoutcovid_France_Youngs = globals()["compar_deaths_France_Youngs"].loc[globals()["covid_deaths_France_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_covid_France_Adults = globals()["covid_deaths_France_Adults"].loc[globals()["covid_deaths_France_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_withoutcovid_France_Adults = globals()["compar_deaths_France_Adults"].loc[globals()["covid_deaths_France_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_covid_France_Elderly = globals()["covid_deaths_France_Elderly"].loc[globals()["covid_deaths_France_Elderly"]["Year"]==2020]["Elderly"].values[0]
Deaths_withoutcovid_France_Elderly = globals()["compar_deaths_France_Elderly"].loc[globals()["covid_deaths_France_Elderly"]["Year"]==2020]["Elderly"].values[0]


mortality_covid_France_Youngs = globals()["covid_mortality_rate_France_Youngs"].loc[globals()["covid_mortality_rate_France_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_withoutcovid_France_Youngs = globals()["mortality_rate_France_Youngs"].loc[globals()["mortality_rate_France_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_covid_France_Adults = globals()["covid_mortality_rate_France_Adults"].loc[globals()["covid_mortality_rate_France_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_withoutcovid_France_Adults = globals()["mortality_rate_France_Adults"].loc[globals()["mortality_rate_France_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_covid_France_Elderly = globals()["covid_mortality_rate_France_Elderly"].loc[globals()["covid_mortality_rate_France_Elderly"]["Year"]==2020]["Elderly"].values[0]
mortality_withoutcovid_France_Elderly = globals()["mortality_rate_France_Elderly"].loc[globals()["mortality_rate_France_Elderly"]["Year"]==2020]["Elderly"].values[0]











##############################################################################
######################      GERMANY     ######################################
##############################################################################




#SIR_MODEL(80000000,1/3.0,1/7,0.15,180,1000,35,0.5,0.6)
#SIR_MODEL(80000000,1/3.0,1/7,0.15,360,1000,105,0.5,0.6))


contact_rate_Germany = 1/3.0

total_population_Germany_2020 = 0

def funct_totalpopulationGermany():
    for ages in ["Youngs","Adults","Elderly"]:
        globals()["Deaths_2020_"+str(ages)] =\
             globals()["compar_population_Germany_"+str(ages)][ages].loc[globals()\
             ["compar_population_Germany_"+str(ages)]["Year"]==2020]
    total_population = globals()["Deaths_2020_Youngs"]\
                       + globals()["Deaths_2020_Adults"] \
                       + globals()["Deaths_2020_Elderly"]
    return total_population


Population_Germany_2020 = funct_totalpopulationGermany()

total_deaths_Germany_2020= SIR_MODEL('Germany',Population_Germany_2020,\
                                     contact_rate_Germany,\
                                     1/7,0.15,360,1000,105,0.5,0.6)


nr_deaths_Germany_Youngs = rate_Youngs *  total_deaths_Germany_2020
nr_deaths_Germany_Adults = rate_Adults * total_deaths_Germany_2020
nr_deaths_Germany_Elderly = rate_Elderly * total_deaths_Germany_2020



## We make a copy from the number of deaths and we assume that the only year 
## that it changes is in 2020 the rest stays constant 

for ages in ["Youngs","Adults","Elderly"]:
    globals()["covid_deaths_Germany_"+str(ages)] = \
        globals()["compar_deaths_Germany_" + str(ages)].copy()
        
    globals()["covid_population_Germany_"+str(ages)] = \
        globals()["compar_population_Germany_" + str(ages)].copy()

    
    globals()["covid_deaths_Germany_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Germany_"+str(ages)]["Year"]==2020] =\
             globals()["covid_deaths_Germany_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Germany_"+str(ages)]["Year"]==2020]  \
             + globals()["nr_deaths_Germany_"+ str(ages)]
            
            
    globals()["covid_population_Germany_"+str(ages)][ages].loc[globals()\
             ["covid_population_Germany_"+str(ages)]["Year"]==2020] =\
        globals()["covid_population_Germany_"+str(ages)][ages].loc[globals()\
        ["covid_population_Germany_"+str(ages)]["Year"]==2020] +\
            globals()["nr_deaths_Germany_"+ str(ages)]
    
    
    year_Germany = globals()["covid_population_Germany_Youngs"]["Year"]
    
    globals()["mortality_Germany_" + str(ages)] = pd.DataFrame( \
                     (globals()["covid_deaths_Germany_"+str(ages)][ages])/\
                    (globals()["covid_population_Germany_"+str(ages)][ages]))
    
    globals()["covid_mortality_rate_Germany_" + str(ages)] = \
                     pd.concat([year_Germany,globals()["mortality_Germany_" +\
                                                       str(ages)]],axis = 1)
    del year_Germany
    del globals()["mortality_Germany_" + str(ages)]



Deaths_covid_Germany_Youngs = globals()["covid_deaths_Germany_Youngs"].loc[globals()["covid_deaths_Germany_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_withoutcovid_Germany_Youngs = globals()["compar_deaths_Germany_Youngs"].loc[globals()["covid_deaths_Germany_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_covid_Germany_Adults = globals()["covid_deaths_Germany_Adults"].loc[globals()["covid_deaths_Germany_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_withoutcovid_Germany_Adults = globals()["compar_deaths_Germany_Adults"].loc[globals()["covid_deaths_Germany_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_covid_Germany_Elderly = globals()["covid_deaths_Germany_Elderly"].loc[globals()["covid_deaths_Germany_Elderly"]["Year"]==2020]["Elderly"].values[0]
Deaths_withoutcovid_Germany_Elderly = globals()["compar_deaths_Germany_Elderly"].loc[globals()["covid_deaths_Germany_Elderly"]["Year"]==2020]["Elderly"].values[0]


mortality_covid_Germany_Youngs = globals()["covid_mortality_rate_Germany_Youngs"].loc[globals()["covid_mortality_rate_Germany_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_withoutcovid_Germany_Youngs = globals()["mortality_rate_Germany_Youngs"].loc[globals()["mortality_rate_Germany_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_covid_Germany_Adults = globals()["covid_mortality_rate_Germany_Adults"].loc[globals()["covid_mortality_rate_Germany_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_withoutcovid_Germany_Adults = globals()["mortality_rate_Germany_Adults"].loc[globals()["mortality_rate_Germany_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_covid_Germany_Elderly = globals()["covid_mortality_rate_Germany_Elderly"].loc[globals()["covid_mortality_rate_Germany_Elderly"]["Year"]==2020]["Elderly"].values[0]
mortality_withoutcovid_Germany_Elderly = globals()["mortality_rate_Germany_Elderly"].loc[globals()["mortality_rate_Germany_Elderly"]["Year"]==2020]["Elderly"].values[0]






##############################################################################
######################      ITALY     ########################################
##############################################################################




#SIR_MODEL(60050000,1/2.87,1/7,0.15,180,1000,35,0.5,0.6)
#SIR_MODEL(60000000,1/2.87,1/7,0.15,360,1000,105,0.5,0.6)


contact_rate_Italy = 1/2.87

total_population_Italy_2020 = 0

def funct_totalpopulationItaly():
    for ages in ["Youngs","Adults","Elderly"]:
        globals()["Deaths_2020_"+str(ages)] =\
             globals()["compar_population_Italy_"+str(ages)][ages].loc[globals()\
             ["compar_population_Italy_"+str(ages)]["Year"]==2020]
    total_population = globals()["Deaths_2020_Youngs"]\
                       + globals()["Deaths_2020_Adults"] \
                       + globals()["Deaths_2020_Elderly"]
    return total_population


Population_Italy_2020 = funct_totalpopulationItaly()

total_deaths_Italy_2020= SIR_MODEL('Italy',Population_Italy_2020,\
                                     contact_rate_Italy,\
                                     1/7,0.15,360,1000,105,0.5,0.6)


nr_deaths_Italy_Youngs = rate_Youngs *  total_deaths_Italy_2020
nr_deaths_Italy_Adults = rate_Adults * total_deaths_Italy_2020
nr_deaths_Italy_Elderly = rate_Elderly * total_deaths_Italy_2020



## We make a copy from the number of deaths and we assume that the only year 
## that it changes is in 2020 the rest stays constant 

for ages in ["Youngs","Adults","Elderly"]:
    globals()["covid_deaths_Italy_"+str(ages)] = \
        globals()["compar_deaths_Italy_" + str(ages)].copy()
        
    globals()["covid_population_Italy_"+str(ages)] = \
        globals()["compar_population_Italy_" + str(ages)].copy()

    
    globals()["covid_deaths_Italy_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Italy_"+str(ages)]["Year"]==2020] =\
             globals()["covid_deaths_Italy_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Italy_"+str(ages)]["Year"]==2020]  \
             + globals()["nr_deaths_Italy_"+ str(ages)]
            
            
    globals()["covid_population_Italy_"+str(ages)][ages].loc[globals()\
             ["covid_population_Italy_"+str(ages)]["Year"]==2020] =\
        globals()["covid_population_Italy_"+str(ages)][ages].loc[globals()\
        ["covid_population_Italy_"+str(ages)]["Year"]==2020] +\
            globals()["nr_deaths_Italy_"+ str(ages)]
    
    
    year_Italy = globals()["covid_population_Italy_Youngs"]["Year"]
    
    globals()["mortality_Italy_" + str(ages)] = pd.DataFrame( \
                     (globals()["covid_deaths_Italy_"+str(ages)][ages])/\
                    (globals()["covid_population_Italy_"+str(ages)][ages]))
    
    globals()["covid_mortality_rate_Italy_" + str(ages)] = \
                     pd.concat([year_Italy,globals()["mortality_Italy_" +\
                                                       str(ages)]],axis = 1)
    del year_Italy
    del globals()["mortality_Italy_" + str(ages)]



Deaths_covid_Italy_Youngs = globals()["covid_deaths_Italy_Youngs"].loc[globals()["covid_deaths_Italy_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_withoutcovid_Italy_Youngs = globals()["compar_deaths_Italy_Youngs"].loc[globals()["covid_deaths_Italy_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_covid_Italy_Adults = globals()["covid_deaths_Italy_Adults"].loc[globals()["covid_deaths_Italy_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_withoutcovid_Italy_Adults = globals()["compar_deaths_Italy_Adults"].loc[globals()["covid_deaths_Italy_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_covid_Italy_Elderly = globals()["covid_deaths_Italy_Elderly"].loc[globals()["covid_deaths_Italy_Elderly"]["Year"]==2020]["Elderly"].values[0]
Deaths_withoutcovid_Italy_Elderly = globals()["compar_deaths_Italy_Elderly"].loc[globals()["covid_deaths_Italy_Elderly"]["Year"]==2020]["Elderly"].values[0]


mortality_covid_Italy_Youngs = globals()["covid_mortality_rate_Italy_Youngs"].loc[globals()["covid_mortality_rate_Italy_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_withoutcovid_Italy_Youngs = globals()["mortality_rate_Italy_Youngs"].loc[globals()["mortality_rate_Italy_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_covid_Italy_Adults = globals()["covid_mortality_rate_Italy_Adults"].loc[globals()["covid_mortality_rate_Italy_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_withoutcovid_Italy_Adults = globals()["mortality_rate_Italy_Adults"].loc[globals()["mortality_rate_Italy_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_covid_Italy_Elderly = globals()["covid_mortality_rate_Italy_Elderly"].loc[globals()["covid_mortality_rate_Italy_Elderly"]["Year"]==2020]["Elderly"].values[0]
mortality_withoutcovid_Italy_Elderly = globals()["mortality_rate_Italy_Elderly"].loc[globals()["mortality_rate_Italy_Elderly"]["Year"]==2020]["Elderly"].values[0]










##############################################################################
######################      Switzerland     ######################################
##############################################################################




#SIR_MODEL(8050000,1/3.17,1/7,0.15,180,1000,35,0.5,0.6)
#SIR_MODEL(8050000,1/3.17,1/7,0.15,360,1000,35,0.5,0.6)

contact_rate_Switzerland = 1/3.17

total_population_Switzerland_2020 = 0

def funct_totalpopulationSwitzerland():
    for ages in ["Youngs","Adults","Elderly"]:
        globals()["Deaths_2020_"+str(ages)] =\
             globals()["compar_population_Switzerland_"+str(ages)][ages].loc[globals()\
             ["compar_population_Switzerland_"+str(ages)]["Year"]==2020]
    total_population = globals()["Deaths_2020_Youngs"]\
                       + globals()["Deaths_2020_Adults"] \
                       + globals()["Deaths_2020_Elderly"]
    return total_population


Population_Switzerland_2020 = funct_totalpopulationSwitzerland()

total_deaths_Switzerland_2020= SIR_MODEL('Switzerland',Population_Switzerland_2020,\
                                     contact_rate_Switzerland,\
                                     1/7,0.15,360,1000,105,0.5,0.6)


nr_deaths_Switzerland_Youngs = rate_Youngs *  total_deaths_Switzerland_2020
nr_deaths_Switzerland_Adults = rate_Adults * total_deaths_Switzerland_2020
nr_deaths_Switzerland_Elderly = rate_Elderly * total_deaths_Switzerland_2020



## We make a copy from the number of deaths and we assume that the only year 
## that it changes is in 2020 the rest stays constant 

for ages in ["Youngs","Adults","Elderly"]:
    globals()["covid_deaths_Switzerland_"+str(ages)] = \
        globals()["compar_deaths_Switzerland_" + str(ages)].copy()
        
    globals()["covid_population_Switzerland_"+str(ages)] = \
        globals()["compar_population_Switzerland_" + str(ages)].copy()

    
    globals()["covid_deaths_Switzerland_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Switzerland_"+str(ages)]["Year"]==2020] =\
             globals()["covid_deaths_Switzerland_"+str(ages)][ages].loc[globals()\
             ["covid_deaths_Switzerland_"+str(ages)]["Year"]==2020]  \
             + globals()["nr_deaths_Switzerland_"+ str(ages)]
            
            
    globals()["covid_population_Switzerland_"+str(ages)][ages].loc[globals()\
             ["covid_population_Switzerland_"+str(ages)]["Year"]==2020] =\
        globals()["covid_population_Switzerland_"+str(ages)][ages].loc[globals()\
        ["covid_population_Switzerland_"+str(ages)]["Year"]==2020] +\
            globals()["nr_deaths_Switzerland_"+ str(ages)]
    
    
    year_Switzerland = globals()["covid_population_Switzerland_Youngs"]["Year"]
    
    globals()["mortality_Switzerland_" + str(ages)] = pd.DataFrame( \
                     (globals()["covid_deaths_Switzerland_"+str(ages)][ages])/\
                    (globals()["covid_population_Switzerland_"+str(ages)][ages]))
    
    globals()["covid_mortality_rate_Switzerland_" + str(ages)] = \
                     pd.concat([year_Switzerland,globals()["mortality_Switzerland_" +\
                                                       str(ages)]],axis = 1)
    del year_Switzerland
    del globals()["mortality_Switzerland_" + str(ages)]



Deaths_covid_Switzerland_Youngs = globals()["covid_deaths_Switzerland_Youngs"].loc[globals()["covid_deaths_Switzerland_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_withoutcovid_Switzerland_Youngs = globals()["compar_deaths_Switzerland_Youngs"].loc[globals()["covid_deaths_Switzerland_Youngs"]["Year"]==2020]["Youngs"].values[0]
Deaths_covid_Switzerland_Adults = globals()["covid_deaths_Switzerland_Adults"].loc[globals()["covid_deaths_Switzerland_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_withoutcovid_Switzerland_Adults = globals()["compar_deaths_Switzerland_Adults"].loc[globals()["covid_deaths_Switzerland_Adults"]["Year"]==2020]["Adults"].values[0]
Deaths_covid_Switzerland_Elderly = globals()["covid_deaths_Switzerland_Elderly"].loc[globals()["covid_deaths_Switzerland_Elderly"]["Year"]==2020]["Elderly"].values[0]
Deaths_withoutcovid_Switzerland_Elderly = globals()["compar_deaths_Switzerland_Elderly"].loc[globals()["covid_deaths_Switzerland_Elderly"]["Year"]==2020]["Elderly"].values[0]


mortality_covid_Switzerland_Youngs = globals()["covid_mortality_rate_Switzerland_Youngs"].loc[globals()["covid_mortality_rate_Switzerland_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_withoutcovid_Switzerland_Youngs = globals()["mortality_rate_Switzerland_Youngs"].loc[globals()["mortality_rate_Switzerland_Youngs"]["Year"]==2020]["Youngs"].values[0]
mortality_covid_Switzerland_Adults = globals()["covid_mortality_rate_Switzerland_Adults"].loc[globals()["covid_mortality_rate_Switzerland_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_withoutcovid_Switzerland_Adults = globals()["mortality_rate_Switzerland_Adults"].loc[globals()["mortality_rate_Switzerland_Adults"]["Year"]==2020]["Adults"].values[0]
mortality_covid_Switzerland_Elderly = globals()["covid_mortality_rate_Switzerland_Elderly"].loc[globals()["covid_mortality_rate_Switzerland_Elderly"]["Year"]==2020]["Elderly"].values[0]
mortality_withoutcovid_Switzerland_Elderly = globals()["mortality_rate_Switzerland_Elderly"].loc[globals()["mortality_rate_Switzerland_Elderly"]["Year"]==2020]["Elderly"].values[0]

# Calculate the increase in mortality
Increase_Belgium_Y = Deaths_covid_Belgium_Youngs/Deaths_withoutcovid_Belgium_Youngs
Increase_Belgium_A = Deaths_covid_Belgium_Adults/Deaths_withoutcovid_Belgium_Adults
Increase_Belgium_E = Deaths_covid_Belgium_Elderly/Deaths_withoutcovid_Belgium_Elderly

Increase_France_Y = Deaths_covid_France_Youngs/Deaths_withoutcovid_France_Youngs
Increase_France_A = Deaths_covid_France_Adults/Deaths_withoutcovid_France_Adults
Increase_France_E = Deaths_covid_France_Elderly/Deaths_withoutcovid_France_Elderly

Increase_Germany_Y = Deaths_covid_Germany_Youngs/Deaths_withoutcovid_Germany_Youngs
Increase_Germany_A = Deaths_covid_Germany_Adults/Deaths_withoutcovid_Germany_Adults
Increase_Germany_E = Deaths_covid_Germany_Elderly/Deaths_withoutcovid_Germany_Elderly

Increase_Italy_Y = Deaths_covid_Italy_Youngs/Deaths_withoutcovid_Italy_Youngs
Increase_Italy_A = Deaths_covid_Italy_Adults/Deaths_withoutcovid_Italy_Adults
Increase_Italy_E = Deaths_covid_Italy_Elderly/Deaths_withoutcovid_Italy_Elderly

Increase_Switzerland_Y = Deaths_covid_Switzerland_Youngs/Deaths_withoutcovid_Switzerland_Youngs
Increase_Switzerland_A = Deaths_covid_Switzerland_Adults/Deaths_withoutcovid_Switzerland_Adults
Increase_Switzerland_E = Deaths_covid_Switzerland_Elderly/Deaths_withoutcovid_Switzerland_Elderly

Covid_Deaths = [[Deaths_covid_Belgium_Youngs,Deaths_covid_Belgium_Adults,Deaths_covid_Belgium_Elderly],[Deaths_covid_France_Youngs
            ,Deaths_covid_France_Adults,Deaths_covid_France_Elderly],[Deaths_covid_Germany_Youngs,Deaths_covid_Germany_Adults,Deaths_covid_Germany_Elderly]
            ,[Deaths_covid_Italy_Youngs,Deaths_covid_Italy_Adults,Deaths_covid_Italy_Elderly],[Deaths_covid_Switzerland_Youngs,Deaths_covid_Switzerland_Adults,Deaths_covid_Switzerland_Elderly]]

NoCovid_Deaths = [[Deaths_withoutcovid_Belgium_Youngs,Deaths_withoutcovid_Belgium_Adults,Deaths_withoutcovid_Belgium_Elderly],[Deaths_withoutcovid_France_Youngs
            ,Deaths_withoutcovid_France_Adults,Deaths_withoutcovid_France_Elderly],[Deaths_withoutcovid_Germany_Youngs,Deaths_withoutcovid_Germany_Adults,Deaths_withoutcovid_Germany_Elderly]
            ,[Deaths_withoutcovid_Italy_Youngs,Deaths_withoutcovid_Italy_Adults,Deaths_withoutcovid_Italy_Elderly],[Deaths_withoutcovid_Switzerland_Youngs,Deaths_withoutcovid_Switzerland_Adults,Deaths_withoutcovid_Switzerland_Elderly]]

Increase = [[Increase_Belgium_Y,Increase_Belgium_A,Increase_Belgium_E],[Increase_France_Y
            ,Increase_France_A,Increase_France_E],[Increase_Germany_Y,Increase_Germany_A,Increase_Germany_E]
            ,[Increase_Italy_Y,Increase_Italy_A,Increase_Italy_E],[Increase_Switzerland_Y,Increase_Switzerland_A,Increase_Switzerland_E]]


# Paste the increase into Excel
filename ="Number_Of_Deaths_Final.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.active 

# copying the cell values from source  
# excel file to destination excel file 
for i in range (0, 5): 
    for j in range (0, 3): 

        ws1.cell(row = i + 2, column = j + 2).value = Increase[i][j]
        ws1.cell(row = i + 2, column = j + 5).value = Covid_Deaths[i][j]
        ws1.cell(row = i + 2, column = j + 8).value = (Covid_Deaths[i][j] - NoCovid_Deaths[i][j])
  
# saving the destination excel file 
wb1.save(str(filename)) 